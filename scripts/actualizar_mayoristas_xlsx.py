#!/usr/bin/env python
"""
Actualiza la hoja 'Mayoristas' de Listado 2025.xlsx según el PDF semanal de Pedraza.

Reglas:
- Pez en PDF Y en Excel (match por nombre común normalizado + talla):
    → actualizar talla + precio (mayorista + fórmula retail)
- Pez que tenía precio en Excel pero YA NO está en PDF:
    → limpiar talla + precio (dejar vacíos)
- Pez en PDF pero NO existe en Excel:
    → agregar fila nueva en la sección de peces, ordenada alfabéticamente

Guarda una copia actualizada en la carpeta de salida (no modifica el original).

Uso:
  python scripts/actualizar_mayoristas_xlsx.py <pdf_pedraza> <xlsx_listado_2025> [xlsx_pedraza_mayorista]

El PDF semanal de Pedraza es suficiente — el parseo extrae ID, nombre científico,
nombre común, talla y precio mayorista. El tercer argumento es OPCIONAL (ya no
necesario en práctica; Pedraza normalmente solo envía PDF) y sirve únicamente como
fallback para enriquecer nombres canónicos en casos extremos.
"""
import re
import sys
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path(r"C:\Users\drami\Downloads\Entre-peces-listados")

# Columnas (1-indexed) en hoja Mayoristas (estructura simplificada)
COL_SCI = 1      # A — Nombre científico
COL_NAME = 2     # B — Nombre común
COL_SIZE = 3     # C — Talla (cm)
COL_PRICE = 4    # D — Precio unitario

# Sección de peces (inclusiva). Fila 3 = encabezados; la sección de peces
# empieza en la 4 y termina antes de que empiece la sección plantas/gambas/productos.
PECES_ROW_START = 4
PECES_ROW_END = 382   # última fila de peces — después plantas, gambas, fertilizantes
PLANTAS_GAMBAS_START = 383
PLANTAS_GAMBAS_END = 394   # r395+ son productos (se mantienen intactos en la hoja)

NOT_PECES_KEYWORDS = ("neocaridin", "gamba", "caracol", "bee shrimp", "snail")

# Aliases (PDF ↔ Excel) — cuando el nombre común difiere entre las dos fuentes
NAME_ALIASES_TO_EXCEL = {
    "arawana plateada": "arawuana plateada",
    "anostomo": "anostomos",
    "barbus rojo": "barbo rojo",
    "bailarina escama de perla": "bailarina escama de perla",
    "golfish ranchu": "golfish ranchun",
    "tiburon cuatro lineas": "tiburon 4 lineas",
}


def normalize(s):
    if not s:
        return ""
    return (
        unicodedata.normalize("NFD", str(s).lower())
        .encode("ascii", "ignore")
        .decode()
        .strip()
    )


def normalize_size(s):
    """Acepta '4 cm', '4', 4.0, '2,5 cm' → devuelve float."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    txt = str(s).lower().replace(",", ".").replace("cm", "").strip()
    try:
        return float(txt)
    except ValueError:
        return None


def retail_price(mayorista):
    m = int(mayorista)
    if m <= 850:
        return m + 50
    if m <= 9000:
        return m + 500
    return m + 1500


def name_key(name):
    n = normalize(name)
    return NAME_ALIASES_TO_EXCEL.get(n, n)


# ---------- Enriquecimiento con Excel Pedraza ----------
def load_pedraza_catalog(xlsx_path):
    """Carga el Excel maestro de Pedraza → dict {id: {sci, name}} por fila P_NNN.
    Hoja 'PECES' del Mayorista-NN.xlsx."""
    if not xlsx_path or not Path(xlsx_path).exists():
        return {}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "PECES" not in wb.sheetnames:
        return {}
    ws = wb["PECES"]
    cat = {}
    duplicates = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10 or len(row) < 8:
            continue
        # [1]=ID, [3]=SCI, [4]=NAME
        id_ = row[1]
        if not id_ or not str(id_).startswith("P_"):
            continue
        sci = row[3]
        name = row[4]
        if not name:
            continue
        key = str(id_).strip()
        if key in cat:
            duplicates.append(key)
            continue  # first-wins (IDs duplicados en Pedraza causan sobrescritura incorrecta)
        cat[key] = {
            "sci": str(sci).strip() if sci else "",
            "name": str(name).strip(),
        }
    if duplicates:
        print(f"  ⚠ Pedraza Excel tiene {len(duplicates)} IDs duplicados (se conserva el primero): {duplicates[:5]}")
    return cat


# ---------- Parseo PDF ----------
TAX_MODIFIERS = {"sp", "sp.", "spp", "spp.", "cf", "cf.", "aff", "aff.", "nov", "nov."}


def split_sci_name(tokens):
    """Dado una lista de tokens entre categoría y 'N cm', separa sci y nombre común.

    Reglas:
    - Saltar tokens UPPERCASE al inicio (residuo de categoría multi-palabra ej. "CICLIDOS AMERICANOS").
    - sci = primer token Title-case + siguiente token all-lowercase (género + especie).
      Si el tercero es modificador taxonómico (sp, spp, etc.), se incluye en sci.
    - nombre común = resto.
    - Si no hay pattern de binomio, todo va a nombre y sci queda vacío.
    """
    # Saltar tokens UPPERCASE residuales de categoría
    while tokens and tokens[0].isupper() and len(tokens[0]) >= 3 and tokens[0].isalpha():
        tokens = tokens[1:]
    if not tokens:
        return "", ""

    def is_genus(t):
        # Título case: empieza en mayúscula + tiene al menos 1 minúscula, o empieza con punto (".Pterygoplichthys")
        clean = t.lstrip(".")
        if not clean:
            return False
        return clean[0].isupper() and any(c.islower() for c in clean)

    def is_species(t):
        return t.isalpha() and t.islower()

    # Patrón 1: Genus + species (+ opcional modificador)
    if len(tokens) >= 2 and is_genus(tokens[0]) and (is_species(tokens[1]) or tokens[1].lower() in TAX_MODIFIERS):
        sci_end = 2
        if len(tokens) >= 3 and tokens[2].lower() in TAX_MODIFIERS:
            sci_end = 3
        sci = " ".join(tokens[:sci_end])
        name = " ".join(tokens[sci_end:])
        return sci, name

    # Patrón 2: solo un token de género (ej "hibrido", "Trichogaster")
    if len(tokens) >= 2 and (is_genus(tokens[0]) or tokens[0].lower() == "hibrido"):
        # Si el segundo es Title case (nombre común), sci es solo el primero
        return tokens[0], " ".join(tokens[1:])

    # Patrón 3: nada parece sci → todo es nombre común
    return "", " ".join(tokens)


def parse_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    items = []
    for line in text.split("\n"):
        if not re.match(r"^P_\d+\b", line):
            continue
        # Match ID y el resto. Extraemos desde el final: $ + price, luego qty, luego "N cm"
        m = re.match(
            r"^(P_\d+)\s+(.+?)\s+(\d+(?:[.,]\d+)?)\s*cm\s+([\d.,]+)\s+\$\s*(.+)$",
            line,
            flags=re.IGNORECASE,
        )
        if not m:
            continue
        id_, middle, size, qty, price = m.groups()
        try:
            qty_n = int(re.sub(r"[^\d]", "", qty) or "0")
            price_n = int(re.sub(r"[^\d]", "", price) or "0")
        except ValueError:
            continue
        if qty_n <= 0:
            continue
        # middle = "CATEGORIA_O_CATEGORIAS sci_words... name_words..."
        # El primer token siempre es parte de la categoría (UPPERCASE).
        # Lo saltamos y dejamos que split_sci_name maneje categorías multi-palabra.
        parts = middle.split()
        if parts and parts[0].isupper():
            parts = parts[1:]  # saltar primera palabra de categoría
        sci, name = split_sci_name(parts)
        size_f = float(size.replace(",", "."))
        if any(k in normalize(name) for k in NOT_PECES_KEYWORDS):
            continue
        items.append(
            {
                "id": id_,
                "sci": sci.strip(),
                "name": name.strip(),
                "size": size_f,
                "qty": qty_n,
                "mayorista": price_n,
            }
        )
    return items


# ---------- Actualizar Excel ----------
def build_sheet_index(ws):
    """Indexa filas de peces por (name_key, size_float) → row_num.
    También retorna un mapa (name_key) → list of row_nums (para filas sin talla)."""
    by_key = {}
    by_name = {}
    for r in range(PECES_ROW_START, PECES_ROW_END + 1):
        name = ws.cell(row=r, column=COL_NAME).value
        size = ws.cell(row=r, column=COL_SIZE).value
        if not name:
            continue
        nk = name_key(name)
        sz = normalize_size(size)
        if sz is not None:
            by_key[(nk, sz)] = r
        by_name.setdefault(nk, []).append(r)
    return by_key, by_name


def read_section(ws, row_start, row_end):
    """Lee una sección como lista de dicts {sci, name, size, price}. Omite filas completamente vacías."""
    out = []
    for r in range(row_start, row_end + 1):
        sci = ws.cell(row=r, column=COL_SCI).value
        name = ws.cell(row=r, column=COL_NAME).value
        size = ws.cell(row=r, column=COL_SIZE).value
        price = ws.cell(row=r, column=COL_PRICE).value
        if sci is None and name is None and size is None and price is None:
            continue
        out.append({"sci": sci, "name": name, "size": size, "price": price})
    return out


def write_section(ws, row_start, data):
    """Escribe una lista de dicts a filas consecutivas desde row_start.
    Omite celdas merged (el valor queda solo en la top-left del merge)."""
    for i, d in enumerate(data):
        r = row_start + i
        for col, key in [(COL_SCI, "sci"), (COL_NAME, "name"), (COL_SIZE, "size"), (COL_PRICE, "price")]:
            try:
                ws.cell(row=r, column=col, value=d.get(key))
            except AttributeError:
                pass
    return row_start + len(data)


def clear_range(ws, row_start, row_end):
    """Limpia valores en cols A-D entre row_start y row_end (inclusivos).
    Omite celdas que forman parte de un MergedCell (no se pueden escribir)."""
    for r in range(row_start, row_end + 1):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            # MergedCell (no la top-left) tiene el atributo value en solo lectura
            try:
                cell.value = None
            except AttributeError:
                pass


def main():
    if len(sys.argv) < 3:
        print("Uso: python actualizar_mayoristas_xlsx.py <pdf_pedraza> <xlsx_listado_2025> [xlsx_pedraza_mayorista]")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    xlsx_path = Path(sys.argv[2])
    pedraza_xlsx = Path(sys.argv[3]) if len(sys.argv) > 3 else None

    if not pdf_path.exists():
        print(f"No existe: {pdf_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"No existe: {xlsx_path}")
        sys.exit(1)

    # Forzar stdout utf-8 tempranamente
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d")
    out_path = OUTPUT_DIR / f"Listado-2025-Actualizado-{ts}.xlsx"

    # Copiar primero para trabajar sobre la copia
    print(f"Copiando archivo original a {out_path.name}")
    shutil.copy2(xlsx_path, out_path)

    print(f"Parseando PDF: {pdf_path.name}")
    pdf_items = parse_pdf(pdf_path)
    print(f"  {len(pdf_items)} peces disponibles en PDF")

    if pedraza_xlsx and pedraza_xlsx.exists():
        print(f"Enriqueciendo nombres con: {pedraza_xlsx.name}")
        catalog = load_pedraza_catalog(pedraza_xlsx)
        enriched = 0
        for it in pdf_items:
            canon = catalog.get(it["id"])
            if canon:
                if canon["sci"]:
                    it["sci"] = canon["sci"]
                if canon["name"]:
                    it["name"] = canon["name"]
                enriched += 1
        print(f"  {enriched}/{len(pdf_items)} enriquecidos con nombres canónicos de Pedraza")

    print(f"Abriendo Excel: {out_path.name}")
    wb = load_workbook(out_path)
    ws = wb["Mayoristas"]

    # Desmerjar merges en peces + plantas/gambas para evitar conflictos al reescribir
    merges_to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if PECES_ROW_START <= mr.min_row <= PLANTAS_GAMBAS_END:
            merges_to_remove.append(str(mr))
    for mr_str in merges_to_remove:
        ws.unmerge_cells(mr_str)
    if merges_to_remove:
        print(f"Desmerjados {len(merges_to_remove)} rangos en peces/plantas/gambas")

    # 1. Leer secciones actuales
    peces_current = read_section(ws, PECES_ROW_START, PECES_ROW_END)
    plantas_gambas = read_section(ws, PLANTAS_GAMBAS_START, PLANTAS_GAMBAS_END)
    print(f"Leído: {len(peces_current)} peces, {len(plantas_gambas)} plantas+gambas")

    # 2. Procesar updates / clears / add new en memoria
    stats = {"updated": 0, "cleared": 0, "created": 0}
    touched_idx = set()

    # Índices sobre peces_current
    idx_by_key = {}     # (name_key, size) → index
    idx_by_name = {}    # name_key → list of indices
    for i, p in enumerate(peces_current):
        if not p.get("name"):
            continue
        nk = name_key(p["name"])
        idx_by_name.setdefault(nk, []).append(i)
        sz = normalize_size(p.get("size"))
        if sz is not None:
            idx_by_key[(nk, sz)] = i

    new_peces = []
    for pez in pdf_items:
        key = (name_key(pez["name"]), pez["size"])
        idx = idx_by_key.get(key)
        if idx is not None:
            peces_current[idx]["size"] = pez["size"]
            peces_current[idx]["price"] = retail_price(pez["mayorista"])
            if not peces_current[idx].get("sci") and pez["sci"]:
                peces_current[idx]["sci"] = pez["sci"]
            touched_idx.add(idx)
            stats["updated"] += 1
            continue
        # Fallback: fila con mismo nombre y sin size/price
        same_name_idxs = idx_by_name.get(name_key(pez["name"]), [])
        empty_idx = next(
            (i for i in same_name_idxs
             if peces_current[i].get("size") is None
             and peces_current[i].get("price") is None
             and i not in touched_idx),
            None,
        )
        if empty_idx is not None:
            peces_current[empty_idx]["size"] = pez["size"]
            peces_current[empty_idx]["price"] = retail_price(pez["mayorista"])
            if not peces_current[empty_idx].get("sci") and pez["sci"]:
                peces_current[empty_idx]["sci"] = pez["sci"]
            touched_idx.add(empty_idx)
            stats["updated"] += 1
            continue
        # No existe → agregar como nuevo
        new_peces.append({
            "sci": pez["sci"],
            "name": pez["name"],
            "size": pez["size"],
            "price": retail_price(pez["mayorista"]),
        })

    # Limpiar peces no tocados que tenían precio
    for i, p in enumerate(peces_current):
        if i in touched_idx:
            continue
        if p.get("size") is not None or p.get("price") is not None:
            peces_current[i]["size"] = None
            peces_current[i]["price"] = None
            stats["cleared"] += 1

    # Agregar nuevos a la lista
    peces_current.extend(new_peces)
    stats["created"] = len(new_peces)

    # 3. Ordenar alfabéticamente por nombre común (y talla secundaria)
    def sort_key(p):
        nm = normalize(p.get("name") or p.get("sci") or "")
        sz = normalize_size(p.get("size")) or 0
        return (nm, sz)
    peces_current.sort(key=sort_key)

    # 4. Limpiar el área antes de reescribir (evita residuos si la lista es más corta)
    projected_end = PECES_ROW_START + len(peces_current) + len(plantas_gambas) - 1
    clear_end = max(PLANTAS_GAMBAS_END, projected_end)

    # Desmerjar cualquier merge dentro del rango de escritura (incluye la parte de productos
    # que vamos a sobrescribir). Los merges más abajo quedan intactos.
    extra_unmerged = 0
    for mr in list(ws.merged_cells.ranges):
        if PECES_ROW_START <= mr.min_row <= clear_end:
            ws.unmerge_cells(str(mr))
            extra_unmerged += 1
    if extra_unmerged:
        print(f"Desmerjados {extra_unmerged} rangos adicionales en zona de reescritura (r4-r{clear_end})")

    clear_range(ws, PECES_ROW_START, clear_end)

    # 5. Escribir peces ordenados
    next_row = write_section(ws, PECES_ROW_START, peces_current)
    # 6. Escribir plantas + gambas después de peces
    next_row = write_section(ws, next_row, plantas_gambas)

    # Resumen
    if new_peces:
        print("\n--- Peces nuevos agregados alfabéticamente ---")
        for p in sorted(new_peces, key=lambda x: normalize(x["name"])):
            print(f"  {str(p['sci'])[:28]:<30} {str(p['name'])[:30]:<32} {p['size']} cm  ${p['price']:,}".replace(',', '.'))

    if next_row - 1 >= 395:
        print(f"\n⚠ ADVERTENCIA: la sección peces+plantas+gambas se extiende hasta r{next_row - 1}.")
        print(f"  Esto sobrescribe parte de la sección de productos (que empezaba en r395).")
        print(f"  Los productos posteriores a r{next_row - 1} siguen intactos.")

    wb.save(out_path)

    print(f"\n=== RESULTADO ===")
    print(f"  Actualizados:   {stats['updated']}")
    print(f"  Limpiados:      {stats['cleared']}")
    print(f"  Nuevos creados: {stats['created']}")
    print(f"\n📁 Archivo: {out_path}")


if __name__ == "__main__":
    main()
