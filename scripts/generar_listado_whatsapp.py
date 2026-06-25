#!/usr/bin/env python
"""
Generar listado filtrado de peces disponibles para enviar por WhatsApp.

Cruza el PDF semanal de Pedraza con el Excel maestro del mismo proveedor
y produce un Excel + PDF con branding de Entre Peces (4 columnas:
nombre científico, nombre común, talla, precio retail).

Totalmente offline: no toca Supabase ni el sitio web.

Uso:
  python scripts/generar_listado_whatsapp.py <ruta_pdf> <ruta_xlsx>
"""
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

# Forzar UTF-8 en stdout (Windows cp1252 falla con emojis)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import pdfplumber
from svglib.svglib import svg2rlg
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ----- Configuración -----
OUTPUT_DIR = Path(r"C:\Users\drami\Downloads\Entre-peces-listados")
ENTRE_PECES_LOGO_SVG = Path(r"C:\Users\drami\projects\Entre-peces\public\logo.svg")
ENTRE_PECES_LOGO_PNG = Path(r"C:\Users\drami\projects\Entre-peces\public\favicon.png")
BRAND_BLUE = "#2563eb"
BRAND_DARK = "#0c2540"

# Aliases para matching entre PDF y Excel
NAME_ALIASES = {
    "arawana plateada": "arawuana plateada",
    "barbus rojo": "barbo rojo",
    "barbus tigre": "barbos tigre",
    "corydora butroy": "corydora wotroi",
    "cucha hypostomo": "cucha hypostomus",
    "golfish ranchu": "golfish ranchun",
}

# Categorías/palabras que NO son peces (excluir del PDF)
NOT_PECES_KEYWORDS = ("neocaridin", "gamba", "caracol", "bee shrimp", "snail")


# ----- Helpers -----
def normalize(s):
    if not s:
        return ""
    return (
        unicodedata.normalize("NFD", str(s).lower())
        .encode("ascii", "ignore")
        .decode()
        .strip()
    )


def normalize_size(s):
    if not s:
        return ""
    return str(s).lower().replace(",", ".").strip()


def retail_price(mayorista):
    m = int(mayorista)
    if m <= 850:
        return m + 50
    if m <= 9000:
        return m + 500
    return m + 1500


def fmt_cop(n):
    return f"${int(n):,}".replace(",", ".")


def key_of(name, size):
    n = normalize(name)
    return (NAME_ALIASES.get(n, n), normalize_size(size))


# ----- Parseo de PDF Pedraza -----
def parse_pdf_ids(pdf_path):
    """Extrae (id, sci, name, size, qty, mayorista) de cada fila del PDF."""
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    rows = []
    for line in text.split("\n"):
        if not re.match(r"^P_\d+\b", line):
            continue
        # Pattern completo: ID CAT SCI... NAME... SIZE cm QTY $ PRICE_PARTS
        m = re.match(
            r"^(P_\d+)\s+(\S+)\s+(.+?)\s+(\d+(?:[.,]\d+)?)\s*cm\s+([\d.,]+)\s+\$\s*(.+)$",
            line,
            flags=re.IGNORECASE,
        )
        if not m:
            continue
        id_, cat, middle, size, qty, price = m.groups()
        try:
            qty_n = int(re.sub(r"[^\d]", "", qty) or "0")
            price_n = int(re.sub(r"[^\d]", "", price) or "0")
        except ValueError:
            continue
        # Heurística: nombre científico es las primeras ~2-3 palabras con mayúscula/lowercase de "middle",
        # el resto es nombre común. Regla simple: primeras 2 palabras como sci, resto como name.
        parts = middle.split()
        if len(parts) >= 3:
            sci = " ".join(parts[:2])
            name = " ".join(parts[2:])
        else:
            sci = ""
            name = middle
        rows.append(
            {
                "id": id_,
                "cat": cat,
                "sci": sci,
                "name": name,
                "size": size.replace(",", ".") + " cm"
                if "cm" not in size.lower()
                else size,
                "qty": qty_n,
                "mayorista": price_n,
            }
        )
    return rows


# ----- Cargar catálogo completo del Excel (fuente de verdad) -----
def load_excel_catalog(xlsx_path):
    """Devuelve dict {id: {sci, name, size, mayorista}} desde la hoja PECES."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["PECES"]
    catalog = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10:
            continue
        # Columnas: [0]=vacía, [1]=ID, [2]=CATEGORIA, [3]=SCI, [4]=NOMBRE, [5]=TALLA, [6]=CANT, [7]=PRECIO
        if len(row) < 8:
            continue
        id_, cat, sci, name, size, stock, price = row[1], row[2], row[3], row[4], row[5], row[6], row[7]
        if not id_ or not name or not size or not price:
            continue
        id_ = str(id_).strip()
        if not id_.startswith("P_"):
            continue
        try:
            mayorista = int(price)
        except (ValueError, TypeError):
            continue
        catalog[id_] = {
            "sci": str(sci).strip() if sci else "",
            "name": str(name).strip(),
            "size": str(size).strip().replace(",", "."),
            "mayorista": mayorista,
        }
    return catalog


def build_items(pdf_rows, catalog):
    """Cruza IDs del PDF con el catálogo Excel (preferente, más canónico).
    Si el ID no está en Excel, usa los datos del PDF como fallback.
    Filtra qty=0 y no-peces."""
    items = []
    excel_missing = 0
    for r in pdf_rows:
        if r["qty"] <= 0:
            continue
        # PDF es fuente de verdad para talla + mayorista (datos semanales reales).
        # Excel se usa solo para enriquecer nombre común y científico (más canónicos).
        cat_item = catalog.get(r["id"])
        if cat_item:
            sci = cat_item["sci"] or r["sci"]
            name = cat_item["name"] or r["name"]
        else:
            excel_missing += 1
            sci = r["sci"]
            name = r["name"]
        size = r["size"]
        mayorista = r["mayorista"]
        low = normalize(name)
        if any(k in low for k in NOT_PECES_KEYWORDS):
            continue
        items.append(
            {
                "id": r["id"],
                "sci": sci,
                "name": name,
                "size": size,
                "qty": r["qty"],
                "mayorista": mayorista,
            }
        )
    if excel_missing:
        print(f"  {excel_missing} IDs no están en Excel (usando datos del PDF directo)")
    return items


# ----- Generar Excel con branding Entre Peces -----
def generate_excel(items, output_path, logo_png):
    wb = Workbook()
    ws = wb.active
    ws.title = "Disponibilidad"

    # Logo (fila 1-3, columna A)
    if logo_png.exists():
        img = XLImage(str(logo_png))
        img.width = 60
        img.height = 60
        img.anchor = "A1"
        ws.add_image(img)

    # Título
    ws["B1"] = "ENTRE PECES"
    ws["B1"].font = Font(name="Calibri", size=20, bold=True, color="2563EB")
    ws["B2"] = "Listado Disponible"
    ws["B2"].font = Font(name="Calibri", size=13, color="0C2540")
    ws["B3"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    # Encabezados tabla en fila 5
    headers = ["Nombre científico", "Nombre común", "Talla", "Precio (COP)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="2563EB")
        c.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[5]:
        c.border = border

    # Filas de datos
    for i, it in enumerate(items, 6):
        ws.cell(row=i, column=1, value=it["sci"]).font = Font(italic=True, size=10)
        ws.cell(row=i, column=2, value=it["name"]).font = Font(size=10)
        c3 = ws.cell(row=i, column=3, value=it["size"])
        c3.alignment = Alignment(horizontal="center")
        c3.font = Font(size=10)
        c4 = ws.cell(row=i, column=4, value=retail_price(it["mayorista"]))
        c4.number_format = '"$"#,##0'
        c4.alignment = Alignment(horizontal="right")
        c4.font = Font(size=10, bold=True, color="0C2540")
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = border
        # zebra
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="F5F7FA")

    # Anchos
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16

    # Pie
    foot_row = len(items) + 7
    ws.cell(
        row=foot_row,
        column=1,
        value="Entre Peces · WhatsApp +57 312 438 0879 · Envío gratis desde $200.000",
    ).font = Font(size=9, italic=True, color="666666")
    ws.merge_cells(
        start_row=foot_row, start_column=1, end_row=foot_row, end_column=4
    )
    ws.cell(row=foot_row, column=1).alignment = Alignment(horizontal="center")

    wb.save(output_path)


# ----- Generar PDF con branding Entre Peces -----
def generate_pdf(items, output_path, logo_png):
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        title="Entre Peces - Listado Disponible",
        author="Entre Peces",
    )
    styles = getSampleStyleSheet()
    brand_blue = colors.HexColor(BRAND_BLUE)
    brand_dark = colors.HexColor(BRAND_DARK)
    story = []

    # Encabezado con logo
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=22,
        textColor=brand_blue,
        alignment=TA_CENTER,
        spaceAfter=2,
        leading=24,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=11,
        alignment=TA_CENTER,
        textColor=brand_dark,
        spaceAfter=2,
    )
    date_style = ParagraphStyle(
        "Date",
        parent=styles["Normal"],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )

    # Logo vectorial desde SVG (nítido en cualquier escala)
    if ENTRE_PECES_LOGO_SVG.exists():
        drawing = svg2rlg(str(ENTRE_PECES_LOGO_SVG))
        target = 2.2 * cm
        scale = target / drawing.width
        drawing.width *= scale
        drawing.height *= scale
        drawing.scale(scale, scale)
        drawing.hAlign = "CENTER"
        story.append(drawing)
    story.append(Spacer(1, 0.1 * cm))
    story.append(Paragraph("ENTRE PECES", title_style))
    story.append(Paragraph("Listado Disponible", sub_style))
    meses_es = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    ]
    _today = datetime.now()
    fecha_es = f"{_today.day} de {meses_es[_today.month - 1]} de {_today.year}"
    story.append(Paragraph(fecha_es, date_style))
    story.append(Spacer(1, 0.5 * cm))

    # Tabla
    italic_sci = ParagraphStyle(
        "sci",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=9,
        textColor=colors.HexColor("#444444"),
        leading=11,
    )
    name_style = ParagraphStyle(
        "name",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        textColor=brand_dark,
        leading=11,
    )

    data = [["Nombre científico", "Nombre común", "Talla", "Precio"]]
    for it in items:
        data.append(
            [
                Paragraph(it["sci"] or "—", italic_sci),
                Paragraph(it["name"], name_style),
                it["size"],
                fmt_cop(retail_price(it["mayorista"])),
            ]
        )

    col_widths = [6 * cm, 6.5 * cm, 2 * cm, 3 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), brand_blue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (2, 1), (2, -1), "CENTER"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F5F7FA")],
                ),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                (
                    "FONTNAME",
                    (3, 1),
                    (3, -1),
                    "Helvetica-Bold",
                ),
                ("TEXTCOLOR", (3, 1), (3, -1), brand_dark),
            ]
        )
    )
    story.append(table)

    # Pie de página
    story.append(Spacer(1, 0.6 * cm))
    foot_style = ParagraphStyle(
        "Foot",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=colors.grey,
        alignment=TA_CENTER,
        leading=11,
    )
    story.append(
        Paragraph(
            "Entre Peces · WhatsApp <b>+57 312 438 0879</b> · Envío gratis desde $200.000",
            foot_style,
        )
    )
    story.append(
        Paragraph(
            f"Total: <b>{len(items)}</b> peces disponibles",
            foot_style,
        )
    )

    doc.build(story)


# ----- Main -----
def main():
    if len(sys.argv) < 3:
        print("Uso: python generar_listado_whatsapp.py <pdf_pedraza> <xlsx_pedraza>")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    xlsx_path = Path(sys.argv[2])
    if not pdf_path.exists():
        print(f"No existe: {pdf_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"No existe: {xlsx_path}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Cargando catálogo Excel: {xlsx_path.name}")
    catalog = load_excel_catalog(xlsx_path)
    print(f"  {len(catalog)} productos en Excel")

    print(f"Parseando PDF: {pdf_path.name}")
    pdf_rows = parse_pdf_ids(pdf_path)
    print(f"  {len(pdf_rows)} filas disponibles en PDF")

    items = build_items(pdf_rows, catalog)
    print(f"  {len(items)} peces finales (tras filtrar no-peces / stock 0 / IDs faltantes)")

    # Orden alfabético por nombre común
    items.sort(key=lambda x: normalize(x["name"]))

    # Para Excel usamos el PNG del favicon; para PDF se usa el SVG directo (vectorial)
    logo_png = ENTRE_PECES_LOGO_PNG

    ts = datetime.now().strftime("%Y%m%d")
    xlsx_out = OUTPUT_DIR / f"Entre-Peces-Disponibilidad-{ts}.xlsx"
    pdf_out = OUTPUT_DIR / f"Entre-Peces-Disponibilidad-{ts}.pdf"

    print(f"Generando Excel: {xlsx_out.name}")
    generate_excel(items, xlsx_out, logo_png)

    print(f"Generando PDF:   {pdf_out.name}")
    generate_pdf(items, pdf_out, logo_png)

    print(f"\n✅ Listado generado con {len(items)} peces")
    print(f"📁 Carpeta: {OUTPUT_DIR}")
    print(f"   • {xlsx_out.name}")
    print(f"   • {pdf_out.name}")


if __name__ == "__main__":
    main()
