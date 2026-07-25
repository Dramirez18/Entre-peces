#!/usr/bin/env python
"""
Genera un PDF corporativo sobrio con el listado semanal disponible de Pedraza,
usando precios retail de Entre Peces.

Input:
  - PDF semanal de Pedraza (disponibilidad + precios mayoristas)
  - OPCIONAL: un xlsx con hoja "Mayoristas" para nombres científicos/comunes
    canónicos. Si no se pasa, se busca el más reciente automáticamente; si no
    hay ninguno, el PDF se genera igual con los nombres crudos del PDF.

Output:
  - PDF con logo Entre Peces, 4 columnas (sci, común, talla, precio retail),
    ordenado alfabéticamente, estilo sobrio/corporativo.

No produce Excel ni toca la BD.

Uso:
  python scripts/generar_pdf_semanal.py <pdf_pedraza> [xlsx_nombres_canonicos]
"""
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ---------- Configuración ----------
OUTPUT_DIR = Path(r"C:\Users\drami\Downloads\Entre-peces-listados")
DOWNLOADS_DIR = Path(r"C:\Users\drami\Downloads")
ENTRE_PECES_LOGO = Path(r"C:\Users\drami\projects\Entre-peces\public\logo entre peces.png")

# Paleta sobria corporativa
NAVY = "#0c2540"       # Azul marino oscuro
ACCENT = "#1d4b8c"     # Azul medio para líneas
GREY_DARK = "#3a3a3a"
GREY_MED = "#6b7280"
GREY_LIGHT = "#f3f4f6"

NOT_PECES_KEYWORDS = (
    "neocaridin", "gamba", "caracol", "bee shrimp", "snail",
    "camaron", "palaemonetes",
)
TAX_MODIFIERS = {"sp", "sp.", "spp", "spp.", "cf", "cf.", "aff", "aff.", "nov", "nov."}


# ---------- Helpers ----------
def normalize(s):
    if not s:
        return ""
    return (
        unicodedata.normalize("NFD", str(s).lower())
        .encode("ascii", "ignore")
        .decode()
        .strip()
    )


def retail_price(mayorista):
    m = int(mayorista)
    if m <= 850:
        return m + 50
    if m <= 9000:
        return m + 500
    return m + 1500


def fmt_cop(n):
    return f"${int(n):,}".replace(",", ".")


# ---------- Parseo PDF ----------
def split_sci_name(tokens):
    while tokens and tokens[0].isupper() and len(tokens[0]) >= 3 and tokens[0].isalpha():
        tokens = tokens[1:]
    if not tokens:
        return "", ""

    def is_genus(t):
        # Acepta genus con mayúscula inicial (estándar) O todo en minúsculas
        # (algunos PDFs de Pedraza tienen "ancistrus", "hypancistrus" en minúsculas).
        clean = t.lstrip(".")
        if not clean or not clean[0].isalpha():
            return False
        return clean.isalpha() and (clean[0].isupper() or clean.islower())

    def is_species(t):
        return t.isalpha() and t.islower()

    def is_numeric_suffix(t):
        # Detecta sufijos como "2" en "Hemiancistrus sp. 2" o similares
        return t.isdigit() or re.match(r"^\d+$", t)

    if len(tokens) >= 2 and is_genus(tokens[0]) and (
        is_species(tokens[1]) or tokens[1].lower() in TAX_MODIFIERS
    ):
        sci_end = 2
        # Incluir modificador taxonómico como tercera palabra si aplica
        if len(tokens) >= 3 and tokens[2].lower() in TAX_MODIFIERS:
            sci_end = 3
        # Incluir sufijo numérico después de "sp/sp./spp" (ej. "sp. 2")
        if (
            len(tokens) >= sci_end + 1
            and tokens[sci_end - 1].lower() in TAX_MODIFIERS
            and is_numeric_suffix(tokens[sci_end])
        ):
            sci_end += 1
        return " ".join(tokens[:sci_end]), " ".join(tokens[sci_end:])

    if len(tokens) >= 2 and (is_genus(tokens[0]) or tokens[0].lower() == "hibrido"):
        return tokens[0], " ".join(tokens[1:])

    return "", " ".join(tokens)


def parse_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    items = []
    for line in text.split("\n"):
        if not re.match(r"^P_\d+\b", line):
            continue
        m = re.match(
            r"^(P_\d+)\s+(.+?)\s+(\d+(?:[.,]\d+)?)\s*cm\s+([\d.,]+)\s+\$\s*(.+)$",
            line,
            flags=re.IGNORECASE,
        )
        if not m:
            continue
        id_, middle, size, qty, price = m.groups()
        try:
            qty_n = int(re.sub(r"[^\d]", "", qty) or "0")
            price_n = int(re.sub(r"[^\d]", "", price) or "0")
        except ValueError:
            continue
        if qty_n <= 0:
            continue
        parts = middle.split()
        if parts and parts[0].isupper():
            parts = parts[1:]
        sci, name = split_sci_name(parts)
        if any(k in normalize(name) for k in NOT_PECES_KEYWORDS):
            continue
        items.append({
            "sci": sci.strip(),
            "name": name.strip(),
            "size": float(size.replace(",", ".")),
            "qty": qty_n,
            "mayorista": price_n,
        })
    return items


# ---------- Enriquecer con Listado 2025 ----------
def find_canon_xlsx():
    """Busca el xlsx de nombres canónicos más reciente (hoja Mayoristas).
    Devuelve None si no encuentra ninguno — el PDF se genera igual."""
    candidates = []
    for folder in (OUTPUT_DIR, DOWNLOADS_DIR):
        if folder.exists():
            candidates.extend(folder.glob("Listado*.xlsx"))
    candidates = [c for c in candidates if not c.name.startswith("~$")]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def build_canon_index(xlsx_path):
    """Indexa nombres canónicos desde la hoja Mayoristas de un Listado *.xlsx.
    Key = normalize(name común). Value = {sci, name}."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Mayoristas" not in wb.sheetnames:
        return {}
    ws = wb["Mayoristas"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # saltar encabezados
            continue
        if len(row) < 4:
            continue
        sci = row[0]
        name = row[1]
        if not name:
            continue
        key = normalize(name)
        if key in out:
            continue  # first-wins
        out[key] = {
            "sci": str(sci).strip() if sci else "",
            "name": str(name).strip(),
        }
    return out


def enrich(items, canon):
    enriched = 0
    for it in items:
        k = normalize(it["name"])
        c = canon.get(k)
        if c:
            if c["sci"]:
                it["sci"] = c["sci"]
            it["name"] = c["name"]
            enriched += 1
    return enriched


# ---------- Generar PDF corporativo ----------
def generate_pdf(items, output_path):
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        title="Entre Peces - Listado Disponible",
        author="Entre Peces",
    )
    styles = getSampleStyleSheet()
    navy = colors.HexColor(NAVY)
    accent = colors.HexColor(ACCENT)
    grey_dark = colors.HexColor(GREY_DARK)
    grey_med = colors.HexColor(GREY_MED)
    grey_light = colors.HexColor(GREY_LIGHT)

    story = []

    # ---- Encabezado sobrio ----
    # Logo pequeño centrado (PNG real de Entre Peces)
    if ENTRE_PECES_LOGO.exists():
        logo = RLImage(str(ENTRE_PECES_LOGO), width=2.0 * cm, height=2.0 * cm)
        logo.hAlign = "CENTER"
        story.append(logo)
    story.append(Spacer(1, 0.15 * cm))

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=navy,
        alignment=TA_CENTER,
        leading=18,
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=grey_dark,
        alignment=TA_CENTER,
        leading=12,
        spaceAfter=2,
    )
    date_style = ParagraphStyle(
        "Date",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        textColor=grey_med,
        alignment=TA_CENTER,
    )

    meses = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    ]
    today = datetime.now()
    fecha = f"{today.day} de {meses[today.month - 1]} de {today.year}"

    story.append(Paragraph("ENTRE PECES", title_style))
    story.append(Paragraph("Listado de Disponibilidad", sub_style))
    story.append(Paragraph(fecha, date_style))
    story.append(Spacer(1, 0.4 * cm))

    # ---- Términos y condiciones / aviso mayorista ----
    terms_style = ParagraphStyle(
        "Terms",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        textColor=grey_dark,
        alignment=TA_CENTER,
        leading=10,
        borderPadding=(6, 8, 6, 8),
        borderColor=accent,
        borderWidth=0.5,
        backColor=grey_light,
    )
    terms_text = (
        "<b>SOMOS ENTRE PECES</b> ubicados en la ciudad de Bogotá. "
        "Recuerde enviar el o los videos de aquellos peces muertos en la recepción "
        "del transporte para reposición — sin estos no se procederá. "
        "<b>Pedido mínimo para mayoristas: $200.000 pesos.</b>"
    )
    story.append(Paragraph(terms_text, terms_style))
    story.append(Spacer(1, 0.4 * cm))

    # ---- Tabla ----
    sci_style = ParagraphStyle(
        "sci",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=8.5,
        textColor=grey_med,
        leading=10,
    )
    name_style = ParagraphStyle(
        "name",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=navy,
        leading=11,
    )

    data = [["Nombre científico", "Nombre común", "Talla", "Precio"]]
    for it in items:
        data.append([
            Paragraph(it["sci"] or "—", sci_style),
            Paragraph(it["name"], name_style),
            f"{int(it['size']) if it['size'] == int(it['size']) else it['size']} cm",
            fmt_cop(retail_price(it["mayorista"])),
        ])

    col_widths = [6.3 * cm, 6.6 * cm, 2.0 * cm, 2.6 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Data body
        ("ALIGN", (2, 1), (2, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ("FONTNAME", (3, 1), (3, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (3, 1), (3, -1), navy),
        # Grid discreta
        ("LINEBELOW", (0, 0), (-1, 0), 0.7, navy),
        ("LINEBELOW", (0, 1), (-1, -2), 0.25, colors.HexColor("#d1d5db")),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, navy),
        # Alternar filas muy sutilmente
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, grey_light]),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)

    # ---- Pie de página ----
    story.append(Spacer(1, 0.6 * cm))
    foot_style = ParagraphStyle(
        "Foot",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=grey_med,
        alignment=TA_CENTER,
        leading=10,
    )
    story.append(Paragraph(
        f"Bogotá · +57 312 438 0879 · Envío gratis desde $200.000",
        foot_style,
    ))
    story.append(Paragraph(
        f"Total: {len(items)} peces disponibles",
        foot_style,
    ))

    # Callback para número de página en footer
    def on_page(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(grey_med)
        canvas.drawRightString(
            letter[0] - 1.6 * cm,
            0.8 * cm,
            f"Página {doc_obj.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)


# ---------- Main ----------
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if len(sys.argv) < 2:
        print("Uso: python generar_pdf_semanal.py <pdf_pedraza> [xlsx_nombres_canonicos]")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    if not pdf_path.exists():
        print(f"No existe: {pdf_path}")
        sys.exit(1)

    # El xlsx es solo fuente de nombres canónicos: opcional y auto-detectable.
    if len(sys.argv) >= 3 and sys.argv[2].strip():
        xlsx_path = Path(sys.argv[2])
        if not xlsx_path.exists():
            print(f"Aviso: no existe {xlsx_path} — busco otro automaticamente.")
            xlsx_path = find_canon_xlsx()
    else:
        xlsx_path = find_canon_xlsx()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Parseando PDF: {pdf_path.name}")
    items = parse_pdf(pdf_path)
    print(f"  {len(items)} peces disponibles")

    if xlsx_path:
        print(f"Enriqueciendo con: {xlsx_path.name}")
        canon = build_canon_index(xlsx_path)
        n = enrich(items, canon)
        print(f"  {n}/{len(items)} enriquecidos con nombres canónicos")
    else:
        print("Sin xlsx de nombres canonicos — uso los nombres del PDF tal cual.")

    # Ordenar alfabéticamente por nombre común
    items.sort(key=lambda x: (normalize(x["name"]), x["size"]))

    ts = datetime.now().strftime("%Y%m%d")
    out_path = OUTPUT_DIR / f"Entre-Peces-Disponibilidad-{ts}.pdf"

    print(f"Generando PDF: {out_path.name}")
    generate_pdf(items, out_path)

    print(f"\n✅ Listo. {len(items)} peces en el PDF.")
    print(f"📁 {out_path}")


if __name__ == "__main__":
    main()
